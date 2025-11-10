import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import sys
import pandas as pd
import yaml
import re
import os
from openpyxl.utils import column_index_from_string
from matplotlib.colors import TwoSlopeNorm
from openpyxl.utils import column_index_from_string
from openpyxl import load_workbook
from BabelViscoFDTD.H5pySimple import SaveToH5py, ReadFromH5py

from TranscranialModeling import BabelIntegrationANNULAR_ARRAY
from BabelViscoFDTD.tools.RayleighAndBHTE import InitCuda,InitOpenCL, InitMetal,ForwardSimple
from scipy.optimize import minimize

from multiprocessing import Process,Queue
import time

import traceback
from skimage import metrics


def read_excel_range(file_path, range_str, sheet_name=0, **kwargs):
    """
    Reads a rectangular Excel range (e.g., 'T20:AJ81') into a DataFrame.

    Parameters:
        file_path (str): Path to the Excel file.
        range_str (str): Excel-style range string (e.g., 'T20:AJ81').
        sheet_name (str or int): Sheet name or index (default is 0).
        **kwargs: Additional arguments passed to `pd.read_excel()`.

    Returns:
        pd.DataFrame: Data from the specified Excel range.
    """
    # Parse range like 'T20:AJ81'
    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', range_str.upper())
    if not match:
        raise ValueError("Invalid range format. Use Excel-style range like 'T20:AJ81'.")

    col_start, row_start, col_end, row_end = match.groups()
    row_start, row_end = int(row_start), int(row_end)
    
    # Compute parameters
    skiprows = row_start - 1
    nrows = row_end - row_start 
    usecols = f"{col_start}:{col_end}"
    print('usecols', usecols)

    # Read the DataFrame
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        skiprows=skiprows,
        nrows=nrows,
        usecols=usecols,
        **kwargs
    )
    return df

def read_excel_range_with_numeric_headers(file_path, range_str, sheet_name=0, **kwargs):
    """
    Reads a rectangular Excel range (e.g., 'T20:AJ81') and assigns numeric headers from the first row of the range.

    Parameters:
        file_path (str): Path to the Excel file.
        range_str (str): Excel-style range string (e.g., 'T20:AJ81').
        sheet_name (str or int): Sheet name or index (default is 0).
        **kwargs: Additional arguments passed to `pd.read_excel()`.

    Returns:
        pd.DataFrame: Data from the specified Excel range with cleaned numeric headers.
    """
    # Parse range
    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', range_str.upper())
    if not match:
        raise ValueError("Invalid range format. Use Excel-style range like 'T20:AJ81'.")

    col_start, row_start, col_end, row_end = match.groups()
    row_start, row_end = int(row_start), int(row_end)
    
    # Column indices (1-based)
    col_start_idx = column_index_from_string(col_start)+1
    col_end_idx = column_index_from_string(col_end)

    # Load header row manually using openpyxl
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name if isinstance(sheet_name, str) else wb.sheetnames[sheet_name]]

    # Read header row cells
    header_cells = [
        ws.cell(row=row_start, column=col_idx).value
        for col_idx in range(col_start_idx, col_end_idx + 1)
    ]
    
    # Optional: Convert to float as pandas do weird things with numneric headers
    headers = []
    for cell in header_cells:
        if isinstance(cell, (int, float)):
            headers.append(np.round(float(cell),1))
        else:
            raise ValueError(f"Non-numeric header value encountered: {cell!r}")

    # Read the data (excluding header row)
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        skiprows=row_start,  # skip header row
        nrows=row_end - row_start,  # only data rows
        usecols=f"{col_start}:{col_end}",
        header=None,
        **kwargs
    )
    
    df.columns = headers
    return df

def MakePlots(infname,x,A,Tx,Frequency,ZDim,Locations,CB,df,deviceName,RealWeight=None,
              bUseRayleighPhase=True,dfPhase=None):
    outdir=os.path.dirname(infname)

    rootpath=outdir+os.sep+'Plots-'
        
    MSEnow=[]
    MSEw=[]

    report={}
    Allbase=[]
    Allexperimental=[]
    Allcorrected=[]
    maxp=0.0
    for l in Locations:
        report[l]=np.array(df[l])
        maxp=np.max([maxp,report[l].max()])
    # for l in Locations:
    #     report[l]/=maxp

    
    
    B=np.ones((A.shape[1],1),np.float32)
    AllFigs=[]
    for n,l in enumerate(Locations):
        if len(Locations)>12:
            div =16
            if n % 16 == 0:
                f,ax=plt.subplots(4,4,figsize=(22/3*2,14/3*2))
                ax=ax.flatten()
                AllFigs.append(f)
        else:
            div =12
            if n == 0:
                f,ax=plt.subplots(3,4,figsize=(22/3*2,14/3*2))
                ax=ax.flatten()
                AllFigs.append(f)
        base=np.abs(A[n*ZDim.shape[0]:(n+1)*ZDim.shape[0],:].dot(B)).flatten()
        corrected=np.abs(A[n*ZDim.shape[0]:(n+1)*ZDim.shape[0],:].dot(x)).flatten()
        experimental=np.array(report[l]).flatten()

        Allbase+=base.tolist()
        Allcorrected+=corrected.tolist()
        Allexperimental+=experimental.tolist()

        # report=df[l]/df[l].max()
        if n % div==0:
            ax[n % div].plot(ZDim*1e3,base,':',label='Uncorrected')
            ax[n % div].plot(ZDim*1e3,experimental,label='Experimental')
            ax[n % div].plot(ZDim*1e3,corrected,label='Corrected')
        else:
            ax[n % div].plot(ZDim*1e3,base,':')
            ax[n % div].plot(ZDim*1e3,experimental)
            ax[n % div].plot(ZDim*1e3,corrected)
        try:
            MSEnow+=((base-experimental)**2).to_list()
            MSEw+=((corrected-experimental)**2).to_list()
        except:
            MSEnow+=((base-experimental)**2).tolist()
            MSEw+=((corrected-experimental)**2).tolist()

        ax[n % div].set_title('TPO = '+str(l) + ' mm',fontsize=10)
        if (n+1) % div == 0 or n==len(Locations)-1:
            f.legend(loc='upper center',bbox_to_anchor=[1.055,0.94])

    Allbase=np.array(Allbase)
    Allexperimental=np.array(Allexperimental)
    Allcorrected=np.array(Allcorrected)
    
    MSEnow=metrics.normalized_root_mse(Allbase,Allexperimental)
    MSEw=metrics.normalized_root_mse(Allcorrected,Allexperimental)
    print('*'*60)
    print('Acoustic axis comparison')
    print('NRMSE non corrected',MSEnow)
    print('NRMSE corrected',MSEw)
    print('NRMSE reduction',1.0 -MSEw/MSEnow)

    SSInow=metrics.structural_similarity(Allbase,Allexperimental,data_range=np.max([Allbase.max(),Allexperimental.max()]),full=True)[0]
    SSIw=metrics.structural_similarity(Allcorrected,Allexperimental,data_range=np.max([Allbase.max(),Allexperimental.max()]),full=True)[0]
    print('SSI non corrected',SSInow)
    print('SSI corrected',SSIw)
    print('SSI improvement',SSIw-SSInow)

    with open(outdir+os.sep+'Part1_Stats.csv','w') as fstat:
        fstat.write('SSI non corrected,SSI corrected,NRMSE non corrected,NRMSE corrected\n')
        fstat.write('%3.2f,%3.2f,%3.2f,%3.2f\n' %(SSInow,SSIw,MSEnow,MSEw))
    
    for n,f in enumerate(AllFigs):
        f.supxlabel('Z (mm)')
        f.supylabel('Pressure (a.u.)')
        f.suptitle('Comparison of corrected and uncorrected source - SSI = %3.2f (uncorrected) and %3.2f (corrected)' %(SSInow,SSIw),fontsize=14)
        f.tight_layout()
        f.savefig(rootpath+'AcProfiles-%02i.pdf' % (n+1),bbox_inches='tight')
 
    def PlotWeight(ax,xd,norm,cmap):
        
        nelem=0
        for VertDisplay, FaceDisplay in zip(Tx['RingVertDisplay'],
                                            Tx['RingFaceDisplay']):
            for e in VertDisplay[FaceDisplay][:,:,:2]:
                color = cmap(norm(xd[nelem]))
                pp3 = plt.Polygon(e*1e3,linewidth=0.1,color=color)
                ax.add_patch(pp3)
                nelem+=1
        ax.set_aspect('equal')
        for din,dout in zip(CB._InDiameters,CB._OutDiameters):
            ax.add_patch(plt.Circle((0,0),din/2*1e3,ls='--',linewidth=1,color='g',fill=False))
            ax.add_patch(plt.Circle((0,0),dout/2*1e3,ls='--',linewidth=1,color='g',fill=False))
        
        ax.set_xlabel('X (mm)')
        ax.set_ylabel('Y (mm)')
        ax.set_xlim(-Tx['Aperture']/2*1e3-2,Tx['Aperture']/2*1e3+2)
        ax.set_ylim(-Tx['Aperture']/2*1e3-2,Tx['Aperture']/2*1e3+2)
        # plt.colorbar(im)
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        sm.set_array([])  # Needed only for colorbar
        plt.colorbar(sm, ax=ax)

    if RealWeight is None:
    
        
        # norm = matplotlib.colors.Normalize(vmin=x.min(), vmax=x.max())  # Normalize scalar between 0 and 1
        if np.iscomplex(x[0]):
            nplots=2  
        else:
            nplots=1
            
        if nplots==1:
            f,axs=plt.subplots(1,nplots,figsize=(5,4))
        else:
            f,axs=plt.subplots(1,nplots,figsize=(12,4))
        if nplots>1:
            axs=axs.flatten()
        else:
            axs=[axs]

        for nf in range(nplots):
            ax=axs[nf]
            if nf==0:
                cmap = plt.cm.jet
                xd=np.abs(x)
                if xd.max()<=1.0:
                    norm =  matplotlib.colors.Normalize(vmin=0.0, vmax=1.0)
                else:
                    norm =  matplotlib.colors.TwoSlopeNorm(vmin=0.0, vmax=xd.max(),vcenter=1.0)
            else:
                cmap = plt.cm.coolwarm
                xd=np.angle(x)
                norm =  matplotlib.colors.TwoSlopeNorm(vmin=-np.pi, vmax=np.pi,vcenter=0.0)

            PlotWeight(ax,xd,norm,cmap)
            if nf==0:
                ax.set_title('Optimized amplitude',fontsize=12)
            else:
                ax.set_title('Optimized phase',fontsize=12)

    else:

        f,axs=plt.subplots(1,2,figsize=(12,4))
        axs=axs.flatten()
        cmap = plt.cm.coolwarm
        norm =  matplotlib.colors.TwoSlopeNorm(vmin=0.0, vmax=RealWeight.max(),vcenter=1.0)
        PlotWeight(axs[0],RealWeight,norm,cmap)
        axs[0].set_title('Real amplitude',fontsize=12)
        PlotWeight(axs[1],x,norm,cmap)
        axs[1].set_title('Optimized amplitude',fontsize=12)


    plt.savefig(rootpath+'weight.pdf',bbox_inches='tight')

    zf=np.array(df.index).astype(np.float32)*1e-3
    Step=1500/Frequency/6
    zf=np.arange(zf.min(),zf.max()+Step,Step)
    xf=np.arange(Tx['center'][:,0].min(),
                 Tx['center'][:,0].max()+Step,Step)
    
    # xf=np.arange(-20e-3,
    #              20e-3+Step,Step)
    yf=np.zeros(1)
    
    yp,xp,zp=np.meshgrid(yf,xf,zf)
    rf=np.hstack((np.reshape(xp,(xp.size,1)),np.reshape(yp,(xp.size,1)), np.reshape(zp,(xp.size,1)))).astype(np.float32)
    
    
    Material=BabelIntegrationANNULAR_ARRAY.Material
    cwvnb_extlay=np.array(2*np.pi*Frequency/Material['Water'][1]+1j*0).astype(np.complex64)

    def DoXZ(inx):
        
        AcPlanes=[]

        bCalcAvgAmpl=True
        
        for l in Locations:
            ds=np.ones((1))*1e-3 #arbitary number
            center=np.zeros((1,3),np.float32)
            center[0,2]=l*1e-3
            u2back=np.zeros(Tx['NumberElems'],np.complex64)
            
            nBase=0
            # print('Locations Tx and center',Tx['center'].min(axis=0),center)
            if bUseRayleighPhase:
                for n in range(Tx['NumberElems']):
                    u0=np.ones(Tx['elemdims'][n][0],np.complex64)
                    SelCenters=Tx['center'][nBase:nBase+Tx['elemdims'][n][0],:].astype(np.float32)
                    SelDs=Tx['ds'][nBase:nBase+Tx['elemdims'][n][0],:].astype(np.float32)
                    u2back[n]=ForwardSimple(cwvnb_extlay,SelCenters,SelDs,u0,center,deviceMetal=deviceName)[0]
                    nBase+=Tx['elemdims'][n][0]
                AllPhi=np.zeros(Tx['NumberElems'])
                for n in range(Tx['NumberElems']):
                    phi=-np.angle(u2back[n])
                    AllPhi[n]=phi
            else:
                AllPhi=np.zeros(Tx['NumberElems'])
                for n in range(Tx['NumberElems']):
                    AllPhi[n]=np.interp(l,dfPhase.index,dfPhase['EL%i Phase' %(n+1)])
            u0=np.zeros((Tx['center'].shape[0],1),np.complex64)
            nBase=0

            GlobalAverage=0.0
            for n in range(Tx['NumberElems']):
                u0[nBase:nBase+Tx['elemdims'][n][0]]=(np.exp(1j*AllPhi[n])).astype(np.complex64)
                if bCalcAvgAmpl:
                    usel=np.abs(inx.reshape((len(inx),1))[nBase:nBase+Tx['elemdims'][n][0]])
                    print('Average (+/- std.) amplitude element %i = %3.2f(%4.3f)' % (n, np.mean(usel), np.std(usel)))
                    GlobalAverage+=np.mean(usel)
                nBase+=Tx['elemdims'][n][0]
            u0*=inx.reshape((len(inx),1))
            if bCalcAvgAmpl:
                print('Global average Tx = %3.2f' % (GlobalAverage/Tx['NumberElems']))
            bCalcAvgAmpl=False #we omly need it once
            acplane=np.abs(ForwardSimple(cwvnb_extlay,Tx['center'].astype(np.float32),Tx['ds'].astype(np.float32),u0,rf,deviceMetal=deviceName))
            acplane=acplane.reshape((len(xf),len(zf)))
            AcPlanes.append(acplane)
        maxPlanes=np.max(np.hstack(AcPlanes))
        print('maxPlanes',maxPlanes)
        for n in range(len(AcPlanes)):
            AcPlanes[n]/=maxPlanes

        return AcPlanes
    AcPlanes=DoXZ(x)

    def PlotXZ(InPlanes):
        AllFigs=[]
        ext=[xf.min()*1e3,xf.max()*1e3,zf.max()*1e3,zf.min()*1e3]

        for l,acplane,n in zip(Locations,InPlanes,range(len(Locations))): 
            if n % 16 == 0:
                f,axs=plt.subplots(2,8,figsize=(16,6))
                axs=axs.flatten()
                AllFigs.append(f)
            ax=axs[n % 16]
            im=ax.imshow(acplane.T,cmap=plt.cm.jet,extent=ext,vmax=1.0,vmin=0)
            ax.set_title('TPO = '+str(l) + ' mm',fontsize=8)

            if n % 8 != 0:
                ax.tick_params(labelleft=False)

            if n < 8:
                ax.tick_params(labelbottom=False)

            ax.tick_params(axis='both', labelsize=9)

        for f in AllFigs:
            plt.figure(f)
            f.subplots_adjust(right=0.83)
            cbar_ax = f.add_axes([0.85, 0.53, 0.015, 0.35])
            cb=f.colorbar(im, cax=cbar_ax)
            cbar_ax.tick_params(labelsize=9)
            plt.suptitle('XZ planes')

        return AllFigs
    
    AllFigs=PlotXZ(AcPlanes)
        
    for n,f in enumerate(AllFigs):
        f.savefig(rootpath+'Acplanes-%02i.pdf' %(n+1),bbox_inches='tight')
        
    AcPlanesReal=[]
    if RealWeight is not None:
        AcPlanesReal=DoXZ(RealWeight)
        PlotXZ(AcPlanesReal)
        plt.suptitle('REAL - XZ planes')
        plt.savefig(rootpath+'REAL-Acplanes.pdf',bbox_inches='tight')
        
        DiffPlanes=[]
        for n in range(len(AcPlanes)):
            DiffPlanes.append(np.abs(AcPlanes[n]-AcPlanesReal[n]))
        PlotXZ(DiffPlanes)
        plt.suptitle('Diff - XZ planes')
        plt.savefig(rootpath+'diff-Acplanes.pdf',bbox_inches='tight')
        AllError=np.abs(np.hstack(AcPlanes)-np.hstack(AcPlanesReal))
        print('Mean error with real planes',np.mean(AllError))
        print('Std error with real planes',np.std(AllError))
        PlanesMeanError=np.mean(AllError)
        PlanesStdError=np.std(AllError)
        PlanesMSE=np.sum(AllError**2)/len(AllError)
    
    plt.close('all')
    if RealWeight is not None:
        return SSInow,SSIw,PlanesMeanError, PlanesStdError,PlanesMSE
    else:
        return SSInow,SSIw

    
def spherical_cap_area(A, R):
    """
    Calculate the surface area of a spherical cap.

    Parameters:
    A (float): Diameter of the circular aperture.
    R (float): Radius of the sphere (focal length).

    Returns:
    float: Surface area of the spherical cap.
    """
    a = A / 2  # aperture radius
    h = R - np.sqrt(R**2 - a**2)
    area = 2 * np.pi * R * h
    return area

def aperture_from_area(S, R):
    """
    Calculate the aperture diameter A of a spherical cap given surface area S and sphere radius R.

    Parameters:
    S (float): Surface area of the spherical cap.
    R (float): Radius of the sphere (focal length).

    Returns:
    float: Aperture diameter A.
    """
    term = R - (S / (2 * np.pi * R))
    a = np.sqrt(R**2 - term**2)
    A = 2 * a
    return A

def equal_area_ring_apertures(S, R, N):
    """
    Divide a spherical cap of area S and radius R into N rings of equal surface area.
    
    Returns aperture diameters (boundaries) for each ring.
    
    Parameters:
    S (float): Total spherical cap surface area.
    R (float): Sphere radius.
    N (int): Number of concentric rings.
    
    Returns:
    List[float]: List of aperture diameters defining ring boundaries (length N).
    """
    delta_h = S / (2 * np.pi * R * N)
    diameters = []
    for i in range(1, N + 1):
        h_i = i * delta_h
        a_i = np.sqrt(R**2 - (R - h_i)**2)
        A_i = 2 * a_i  # aperture diameter
        diameters.append(A_i)
    return np.array(diameters)

def equal_area_ring_apertures_with_inner(S, R, N, A_min):
    """
    Divide a spherical cap (starting from aperture A_min) into N rings of equal surface area.
    
    Parameters:
    S (float): Total surface area of the spherical cap (excluding inner blocked region).
    R (float): Sphere radius.
    N (int): Number of concentric rings.
    A_min (float): Inner blocked aperture diameter.
    
    Returns:
    List[float]: Aperture diameters (outer edges) of the N rings.
    """
    a_min = A_min / 2
    h_min = R - np.sqrt(R**2 - a_min**2)
    
    delta_h = S / (2 * np.pi * R * N)
    
    diameters = []
    for i in range(1, N + 1):
        h_i = h_min + i * delta_h
        a_i = np.sqrt(R**2 - (R - h_i)**2)
        A_i = 2 * a_i
        diameters.append(A_i)
    
    return np.array(diameters)

def RayleighCoeff(Tx,Loc,cwn,u0):
    Distance=np.linalg.norm(Tx['center']-Loc,axis=1)
    coeff=1j*cwn/(2*np.pi)
    coeff=coeff*np.exp(-1j*cwn*Distance)/Distance*Tx['ds'].flatten()*u0.flatten()
    return coeff.astype(np.complex64)

# ----------------------
# Utility Functions
# ----------------------
def compute_data_term(b, A, e):
    Ab = A @ b
    return np.sum((np.abs(Ab) - e)**2)

def compute_data_gradient(b, A, e):
    Ab = A @ b
    abs_Ab = np.abs(Ab)
    safe_abs = np.where(abs_Ab == 0, 1e-12, abs_Ab)
    Ab_conj = np.conj(Ab)

    grad = np.zeros_like(b)
    for k in range(len(b)):
        partial = np.real(A[:, k] * Ab_conj / safe_abs)
        grad[k] = 2 * np.sum((abs_Ab - e) * partial)
    return grad

def phase_objective(b, A, c):
    phi = np.exp(1j * b)        # element-wise e^{i b_j}
    z = A @ phi                 # resulting complex vector
    return np.sum((np.abs(z) - c) ** 2)

def phase_jacobian(b, A, c):
    phi = np.exp(1j * b)        # e^{i b}
    z = A @ phi                 # shape (M,)
    abs_z = np.abs(z)
    safe_abs = np.where(abs_z == 0, 1e-12, abs_z)  # avoid div-by-zero
    z_conj = np.conj(z)

    grad = np.zeros_like(b)
    for k in range(len(b)):
        d_z_dbk = 1j * A[:, k] * phi[k]  # derivative of z wrt b_k
        term = np.real(z_conj / safe_abs * d_z_dbk)
        grad[k] = 2 * np.sum((abs_z - c) * term)
    return grad

# ----------------------
# Regularization Classes
# ----------------------
class Regularizer:
    def __call__(self, b):
        return 0.0

    def gradient(self, b):
        return np.zeros_like(b)

class L2Regularizer(Regularizer):
    def __init__(self, lam):
        self.lam = lam

    def __call__(self, b):
        return self.lam * np.sum(b**2)

    def gradient(self, b):
        return 2 * self.lam * b

class L1Regularizer(Regularizer):
    def __init__(self, lam, delta=1e-6):
        self.lam = lam
        self.delta = delta

    def __call__(self, b):
        return self.lam * np.sum(np.sqrt(b**2 + self.delta))

    def gradient(self, b):
        return self.lam * b / np.sqrt(b**2 + self.delta)

class EntropyRegularizer(Regularizer):
    def __init__(self, lam):
        self.lam = lam

    def __call__(self, b):
        return self.lam * np.sum(b * np.log(np.clip(b, 1e-12, None)))

    def gradient(self, b):
        return self.lam * (np.log(np.clip(b, 1e-12, None)) + 1)

class GroupL2Regularizer(Regularizer):
    def __init__(self, lam, groups):
        self.lam = lam
        self.groups = groups

    def __call__(self, b):
        return self.lam * sum(np.linalg.norm(b[g]) for g in self.groups)

    def gradient(self, b):
        grad = np.zeros_like(b)
        for g in self.groups:
            norm_g = np.linalg.norm(b[g])
            if norm_g > 1e-12:
                grad[g] += self.lam * b[g] / norm_g
        return grad

class GroupHomogeneityRegularizer(Regularizer):
    def __init__(self, lam, group_indices):
        self.lam = lam
        self.group_indices = group_indices  # list of lists of indices

    def __call__(self, b):
        reg = 0.0
        for group in self.group_indices:
            group_b = b[group]
            mean_b = np.mean(group_b)
            reg += np.sum((group_b - mean_b)**2)
        return self.lam * reg

    def gradient(self, b):
        grad = np.zeros_like(b)
        for group in self.group_indices:
            group_b = b[group]
            mean_b = np.mean(group_b)
            for i in group:
                grad[i] += 2 * self.lam * (b[i] - mean_b)
        return grad
    
class TotalVariationRegularizer(Regularizer):
    def __init__(self, lam, delta=1e-6):
        self.lam = lam
        self.delta = delta

    def __call__(self, b):
        diff = np.diff(b)
        return self.lam * np.sum(np.sqrt(diff**2 + self.delta))

    def gradient(self, b):
        grad = np.zeros_like(b)
        diff = np.diff(b)
        denom = np.sqrt(diff**2 + self.delta)
        grad[:-1] -= self.lam * diff / denom
        grad[1:] += self.lam * diff / denom
        return grad

class ElasticNetRegularizer(Regularizer):
    def __init__(self, lam1, lam2, delta=1e-6):
        self.lam1 = lam1
        self.lam2 = lam2
        self.delta = delta

    def __call__(self, b):
        l1 = np.sum(np.sqrt(b**2 + self.delta))
        l2 = np.sum(b**2)
        return self.lam1 * l1 + self.lam2 * l2

    def gradient(self, b):
        l1_grad = b / np.sqrt(b**2 + self.delta)
        l2_grad = 2 * b
        return self.lam1 * l1_grad + self.lam2 * l2_grad

class TikhonovRegularizer(Regularizer):
    def __init__(self, lam, L):
        self.lam = lam
        self.L = L  # L is a matrix

    def __call__(self, b):
        return self.lam * np.sum((self.L @ b)**2)

    def gradient(self, b):
        return 2 * self.lam * self.L.T @ (self.L @ b)
# ----------------------
# Objective with Regularization
# ----------------------
def objective_reg(b, A, e, regularizer):
    return compute_data_term(b, A, e) + regularizer(b)

def jacobian_reg(b, A, e, regularizer):
    return compute_data_gradient(b, A, e) + regularizer.gradient(b)

# ----------------------
# Optimization Wrapper
# ----------------------
def optimize_b(A, e, b0, regularizer,amplitudeLimit=100.0):
    options={'maxfun':2000,'disp':0}
    return minimize(
        objective_reg, b0, args=(A, e, regularizer),
        jac=jacobian_reg,
        bounds=[(1e-6, amplitudeLimit)] * len(b0),
        method='L-BFGS-B',
        options=options
    )

def objective_reg_phaae(b, A, e, regularizer):
    return phase_objective(b, A, e) + regularizer(b)

def jacobian_reg_phase(b, A, e, regularizer):
    return phase_jacobian(b, A, e) + regularizer.gradient(b)

# ----------------------
# Optimization Wrapper
# ----------------------
def optimize_b_phase(A, e, b0, regularizer):
    options={'maxfun':2000,'disp':2}
    return minimize(
        objective_reg_phaae, b0, args=(A, e, regularizer),
        jac=jacobian_reg_phase,
        bounds=[(-np.pi, np.pi)] * len(b0),
        method='L-BFGS-B',
        options=options
    )

def complex_objective(b_complex, A, e,Weights):
    b =b_complex[:A.shape[1]]*np.exp(1j*b_complex[A.shape[1]:])
    Ab = A @ b
    residuals = np.abs(Ab) - e
    return np.sum((residuals ** 2)*Weights)

def complex_gradient(b_complex, A, e,Weights):
    b =b_complex[:A.shape[1]]*np.exp(1j*b_complex[A.shape[1]:])
    Ab = A @ b
    abs_Ab = np.abs(Ab)
    safe_abs = np.where(abs_Ab < 1e-12, 1e-12, abs_Ab)
    z_ratio = ((abs_Ab - e)/ safe_abs) * Ab *Weights # shape (M,)
    grad_complex = 2*(A.conj().T @ z_ratio)       # shape (N,)

    # Return real-valued gradient for real optimizer
        # return  np.hstack([np.abs(grad_complex), np.angle(grad_complex)])
    return np.hstack([grad_complex.real, grad_complex.imag])


def objective_reg_complex(b, A, e, regularizer,Weights):
    return complex_objective(b, A, e,Weights) + regularizer(np.abs(b[:A.shape[1]]*np.exp(1j*b[A.shape[1]:])))

def jacobian_reg_complex(b, A, e, regularizer,Weights):
    regularizationGrad=np.zeros(A.shape[1]*2,dtype=np.float32)
    #we regularize first the amplitude, then the phase
    regularizationGrad[:A.shape[1]]=regularizer.gradient(b[:A.shape[1]])
    regularizationGrad[A.shape[1]:]=regularizer.gradient(b[A.shape[1]:])
    return complex_gradient(b, A, e,Weights) + regularizationGrad



# ----------------------
# Optimization Wrapper
# ----------------------
def optimize_b_complex(A, e, b0, regularizer,display=0,amplitudeLimit=100.0,Weights=1.0):
    options={'maxfun':2000,'disp':display}
    bounds=[(1e-6, amplitudeLimit)] * (len(b0)//2)
    bounds+=[(-np.pi, np.pi)] * (len(b0)//2)
    return minimize(
        objective_reg_complex, b0, args=(A, e, regularizer,Weights),
        jac=jacobian_reg_complex,
        bounds=bounds,
        method='L-BFGS-B',
        options=options)

def complex_objective_RI(b_complex, A, e,Weights):
    N = A.shape[1]
    ml = np
    H=A
    u=b_complex[:N] + 1j * b_complex[N:]
    p_abs=e
    W=Weights
        
    Hu = H @ u
    abs_Hu = ml.abs(Hu)
    eps = 1e-12  # prevent division by zero

    # === Loss ===
    residual = abs_Hu - p_abs
    loss_data = ml.sum((residual**2)*W)
    
    # === Gradient ===
    # ∂/∂u_r and ∂/∂u_i from chain rule
    a_r = ml.real(Hu)
    a_i = ml.imag(Hu)
    abs_safe = ml.maximum(abs_Hu, eps)
    
    factor = 2 * residual / abs_safe * W # shape (M,)
    grad_ur = ml.real(H.conj().T @ (factor * a_r + 1j * factor * a_i))
    grad_ui = ml.imag(H.conj().T @ (factor * a_r + 1j * factor * a_i))

    grad_total = np.concatenate([grad_ur, grad_ui]) #here, if mlx, will be evaluated
    return loss_data, grad_total
    


def objective_reg_complex_RI(b, A, e, regularizer,Weights):
    loss_total, grad_total = complex_objective_RI(b, A, e,Weights)
    loss_total+=regularizer(b[:A.shape[1]])+regularizer(b[A.shape[1]:])
    grad_total[:A.shape[1]]+=regularizer.gradient(b[:A.shape[1]])
    grad_total[A.shape[1]:]+=regularizer.gradient(b[A.shape[1]:])
    return loss_total, grad_total

def optimize_b_complex_RI(A, e, b0, regularizer,display=0,amplitudeLimit=100.0,Weights=1.0):
    options={'maxfun':2000,'disp':display}
    bounds=[(-amplitudeLimit, amplitudeLimit)] * (len(b0)//2)
    bounds+=[(-amplitudeLimit, amplitudeLimit)] * (len(b0)//2)
    return minimize(
        objective_reg_complex_RI, b0, args=(A, e, regularizer,Weights),
        jac=True,
        bounds=bounds,
        method='L-BFGS-B',
        options=options)

def RUN_FITTING(TxConfig,
              YAMLConfigFilename,
              deviceName='M3',
              COMPUTING_BACKEND=3):
    #From Report

    if COMPUTING_BACKEND==1:
        InitCuda(deviceName)
    elif COMPUTING_BACKEND==2:
        InitOpenCL(deviceName)
    elif COMPUTING_BACKEND==3:
        InitMetal(deviceName)

    with open(YAMLConfigFilename, 'r') as f:
        INPUT_PARAMS = yaml.safe_load(f)

    print('Calibration INPUT_PARAMS',INPUT_PARAMS)

    # First we load the mandatory parameters
    dffilename=INPUT_PARAMS['ExcelFileProfiles']
    range_str=INPUT_PARAMS['ExcelRangeProfiles']
    sheet = range_str.split('!')[0] if '!' in range_str else 0
    range_str = range_str.split('!')[1] if '!' in range_str else range_str
    rootnamepath=INPUT_PARAMS['OutputResultsPath']
    lam = INPUT_PARAMS['Lambda']
    Frequency=INPUT_PARAMS['Frequency']

    # we load commonn optional parameters 
    bUseRayleighPhase = INPUT_PARAMS.get('UseRayleighPhase',True)
    dfPhasefilename = INPUT_PARAMS.get('ExcelFilePhase', None)
    rangePhase_str = INPUT_PARAMS.get('ExcelRangePhase', None)
    if rangePhase_str is not None:
        sheetPhase_name = rangePhase_str.split('!')[0] if '!' in rangePhase_str else 0
        rangePhase_str = rangePhase_str.split('!')[1] if '!' in rangePhase_str else rangePhase_str

    # load more experimental parameters
    FitType=INPUT_PARAMS.get('FitType','RealImag')
    regularizer=INPUT_PARAMS.get('Regularizer','Grouped')
    config = INPUT_PARAMS.get('Config',1)
    InnerD = INPUT_PARAMS.get('InnerDiameter', 0.0)
    amplitudeLimit=INPUT_PARAMS.get('AmplitudeLimit', 4.0)

    GeometricFocus=TxConfig['FocalLength']
    FocalDepth=TxConfig['NaturalOutPlaneDistance']
    
    Aperture=TxConfig['TxDiam']
    InD=np.array(TxConfig['InDiameters'])
    OutD=np.array(TxConfig['OutDiameters'])

    if not os.path.isdir(rootnamepath):
        os.makedirs(rootnamepath)
    

    df=read_excel_range_with_numeric_headers(dffilename,range_str,sheet_name=sheet,index_col=0)
    dfPhase=None
    if dfPhasefilename is not None:
        dfPhase=read_excel_range_with_numeric_headers(dfPhasefilename,rangePhase_str,sheet_name=sheetPhase_name,index_col=0)
    
    DistanceZ0Outplane= GeometricFocus-FocalDepth
    print('DistanceZ0Outplane',DistanceZ0Outplane)

    FocalLength=GeometricFocus

    AreaCap=spherical_cap_area(Aperture, FocalLength)
    if InnerD>0:
        AreaCap-=spherical_cap_area(InnerD, FocalLength)
    RadiusIn=aperture_from_area(AreaCap/2, FocalLength)
    print('Aperture',Aperture)
    print('AreaCap',AreaCap)
    print('RadiusIn',RadiusIn)


    if config ==1:
        InDiameters=InD
        OutDiameters=OutD
    else:
        if InnerD>0.0:
            print('using ring calculation with dead inner diameter', InnerD)
            ArtRings=equal_area_ring_apertures_with_inner(AreaCap, FocalLength, len(InD), InnerD)
        else:
            ArtRings=equal_area_ring_apertures(AreaCap, FocalLength, len(InD))
        InDiameters=ArtRings.copy()
        InDiameters[1:]=InDiameters[0:-1]
        InDiameters[0]=InnerD
        OutDiameters=ArtRings

    print('rings areas',(spherical_cap_area(OutDiameters,FocalLength)-spherical_cap_area(InDiameters,FocalLength))*1e6)

    CB=BabelIntegrationANNULAR_ARRAY.SimulationConditions(Frequency=Frequency,
                                                      Aperture=Aperture, # m, aperture of the Tx, used to calculated cross section area entering the domain
                                                      FocalLength=FocalLength,
                                                      InDiameters=InDiameters, #inner diameter of rings
                                                      OutDiameters=OutDiameters)
    Tx=CB.GenTx(PPWSurface=8)
    for k in ['center','RingVertDisplay','elemcenter']:
        if k == 'RingVertDisplay':
            for n in range(len(Tx[k])):
                Tx[k][n][:,2]-=DistanceZ0Outplane
        else:
            Tx[k][:,2]-=DistanceZ0Outplane

    ZDim=np.array(df.index).astype(np.float32)*1e-3
    Step=1500/Frequency/6
    ZDim=np.arange(ZDim.min(),ZDim.max()+Step,Step)
    XDim=np.zeros(1,dtype=np.float32)
    YDim=np.zeros(1,dtype=np.float32)

    yp,xp,zp=np.meshgrid(YDim,XDim,ZDim)
    rf=np.hstack((np.reshape(xp,(len(ZDim),1)),np.reshape(yp,(len(ZDim),1)), np.reshape(zp,(len(ZDim),1)))).astype(np.float32)


    Material=BabelIntegrationANNULAR_ARRAY.Material
    cwvnb_extlay=np.array(2*np.pi*Frequency/Material['Water'][1]+1j*0).astype(np.complex64)

    Locations=np.array(df.columns)

    PhaseReprogram={}

    MaxIntensity=0.0
    MaxIntensityReport=0.0
    IdealAcAxes={}
    for l in Locations:
        ds=np.ones((1))*1e-3 #arbitary number
        center=np.zeros((1,3),np.float32)
        center[0,2]=l*1e-3
        u2back=np.zeros(Tx['NumberElems'],np.complex64)
        nBase=0
        # print('Locations Tx and center',Tx['center'].min(axis=0),center)
        if bUseRayleighPhase:
            for n in range(Tx['NumberElems']):
                u0=np.ones(Tx['elemdims'][n][0],np.complex64)
                SelCenters=Tx['center'][nBase:nBase+Tx['elemdims'][n][0],:].astype(np.float32)
                SelDs=Tx['ds'][nBase:nBase+Tx['elemdims'][n][0],:].astype(np.float32)
                u2back[n]=ForwardSimple(cwvnb_extlay,SelCenters,SelDs,u0,center,deviceMetal=deviceName)[0]
                nBase+=Tx['elemdims'][n][0]
            AllPhi=np.zeros(Tx['NumberElems'])
            for n in range(Tx['NumberElems']):
                phi=-np.angle(u2back[n])
                AllPhi[n]=phi
        else:
            AllPhi=np.zeros(Tx['NumberElems'])
            for n in range(Tx['NumberElems']):
                AllPhi[n]=np.interp(l,dfPhase.index,dfPhase['EL%i Phase' %(n+1)])
        
        # print('AllPhi',np.rad2deg(AllPhi))
        BasePhasedArrayProgramming=np.exp(1j*AllPhi)
        u0=np.zeros((Tx['center'].shape[0],1),np.complex64)
        nBase=0
        for n in range(Tx['NumberElems']):
            u0[nBase:nBase+Tx['elemdims'][n][0]]=(np.exp(1j*AllPhi[n])).astype(np.complex64)
            nBase+=Tx['elemdims'][n][0]

        AcAxis=np.abs(ForwardSimple(cwvnb_extlay,Tx['center'].astype(np.float32),Tx['ds'].astype(np.float32),u0,rf,deviceMetal=deviceName))
        AcAxis=AcAxis**2  # Square the amplitude to get intensity
        IdealAcAxes[l]=AcAxis/AcAxis.max()
        MaxIntensity=np.max([AcAxis.max(),MaxIntensity])
        distance6dB = ZDim[AcAxis/AcAxis.max()>=0.5]*1e3
        length6dB = distance6dB[-1]-distance6dB[0]
        center6dB=np.mean([distance6dB[-1],distance6dB[0]])
        MaxLoc=ZDim[np.argmax(AcAxis)]*1e3
        
        PhaseReprogram[l]=AllPhi

        MaxIntensityReport=np.max([df[l].max(),MaxIntensityReport])


    for l in Locations:
        relative_=df[l].max()/MaxIntensityReport
        IdealAcAxes[l]=IdealAcAxes[l]*relative_

    plt.figure(figsize=(16/3*2,5/3*2))
    lines=[]
    for l in Locations:
        lines.append(plt.plot(np.array(df.index),np.sqrt(df[l]/MaxIntensityReport),'-',label=str(l))[0])
    plt.legend()
    for l,p in zip(Locations,lines):
        plt.plot(ZDim*1e3,np.sqrt(IdealAcAxes[l]),':',color=p.get_color())



    ZDim=np.array(df.index).astype(np.float32)*1e-3
    A=np.zeros((ZDim.shape[0]*len(Locations),Tx['center'].shape[0]),np.complex64)

    for n,l in enumerate(Locations):
        u0=np.zeros((Tx['center'].shape[0],1),np.complex64)
        nBase=0
        for m in range(Tx['NumberElems']):
            u0[nBase:nBase+Tx['elemdims'][m][0]]=(np.exp(1j*PhaseReprogram[l][m])).astype(np.complex64)
            nBase+=Tx['elemdims'][m][0]

        sA=np.zeros((ZDim.shape[0],Tx['center'].shape[0]),np.complex64)
        Loc=np.zeros((1,3))
        for m,z in enumerate(ZDim):
            Loc[0,2]=z
            sA[m,:]=RayleighCoeff(Tx,Loc,cwvnb_extlay,u0)

        B=np.ones((Tx['center'].shape[0],1),np.float32)
        R=np.abs(sA.dot(B))

        # we normalize the amplitude to the maximum of the Rayleigh coefficient and multiplied by the sq. root of maximum of the df[l] (intensity) to match the amplitude
        sA/=R.max() 
        sA*=np.sqrt(df[l].max()/MaxIntensityReport) 

        A[n*ZDim.shape[0]:(n+1)*ZDim.shape[0],:]=sA


    alllines=[]

    plt.figure(figsize=(12,5))
    dfMeasurements=df.copy()
    for n,l in enumerate(Locations):
        dfMeasurements[l]=np.sqrt(dfMeasurements[l]/MaxIntensityReport)
        alllines.append(plt.plot(ZDim, dfMeasurements[l],label=str(l))[0])
    plt.legend()
    for n,l in enumerate(Locations):
        sA=A[n*ZDim.shape[0]:(n+1)*ZDim.shape[0],:]
        plt.plot(ZDim,np.abs(sA.dot(B.flatten())),':',color=alllines[n].get_color())



    sz=len(df[Locations[0]])
    E=np.zeros(sz*len(Locations))
    for n,l in enumerate(Locations):
        E[n*sz:(n+1)*sz]=np.array(dfMeasurements[l])
    

    if regularizer == 'None':
        reg = Regularizer()
    elif regularizer=='L2':

        reg =L2Regularizer(lam=lam)
    elif regularizer in ['Grouped','GroupedL2']:
        groups=[]
        nBase=0
    
        for m in range(Tx['NumberElems']):
            groups.append(np.arange(nBase,nBase+Tx['elemdims'][m][0]))
            nBase+=Tx['elemdims'][m][0]
        if regularizer=='Grouped':
            reg =GroupHomogeneityRegularizer(lam,groups)
        else:
            reg =GroupL2Regularizer(lam,groups)
    else:
        raise ValueError("Unknown regularization " + regularizer)
    if FitType == 'AbsPhase':
            x0=np.zeros(A.shape[1]*2,dtype=np.float32)
            np.random.seed(78) # we use the same seed so we can compare between regularizers
            x0[:A.shape[1]]=np.random.random(A.shape[1]).astype(np.float32)+0.1
            res=optimize_b_complex(A, E, x0,reg,amplitudeLimit=amplitudeLimit)
            res.x=res.x[:A.shape[1]]*np.exp(1j*res.x[A.shape[1]:])
    elif FitType == 'RealImag':
        x0=np.zeros(A.shape[1]*2,dtype=np.float32)
        np.random.seed(78) # we use the same seed so we can compare between regularizers
        x0[:A.shape[1]]=np.random.random(A.shape[1]).astype(np.float32)+0.1
        print('using Real+Imag fitting')
        res=optimize_b_complex_RI(A, E, x0,reg,amplitudeLimit=amplitudeLimit)
        res.x=res.x[:A.shape[1]]+1j*res.x[A.shape[1]:]
    else:
        assert(FitType=='Amp')
        x0=np.ones(A.shape[1])
        res=optimize_b(A, E, x0, reg,amplitudeLimit=amplitudeLimit)


    fname=rootnamepath + os.sep + 'CALIBRATION'

    CalResults = {  'CALIBRATION': res.x,
                    'deviceName':deviceName,
                    'TxConfig':TxConfig,
                    'YAMLConfigFilename': YAMLConfigFilename,
                    'Frequency': Frequency,
                    'FocalLength': FocalLength,
                    'Aperture': Aperture,
                    'InDiameters': InDiameters,
                    'OutDiameters': OutDiameters,
                    'Locations': Locations,
                    'MaxIntensityReport': MaxIntensityReport,
                    'MaxIntensity': MaxIntensity,
                    'Regularizer': regularizer,
                    'bUseRayleighPhase': bUseRayleighPhase}
    
    SaveToH5py(CalResults,fname+'.h5')

    dfMeasurements=df.copy()
    for n,l in enumerate(Locations):
        dfMeasurements[l]=np.sqrt(dfMeasurements[l]/MaxIntensityReport)



    MSENow,MSE_BFGS=MakePlots(fname,res.x,A,Tx,Frequency,ZDim,Locations,CB,dfMeasurements,deviceName,
                                   bUseRayleighPhase=bUseRayleighPhase,dfPhase=dfPhase)
    plt.close('all')
    #we resave the results to add MSE metrics
    CalResults['MSENow']=MSENow
    CalResults['MSE_BFGS']=MSE_BFGS    
    SaveToH5py(CalResults,fname+'.h5')


    return MSENow,MSE_BFGS


def RUN_FITTING_Process(queue,TxConfig, YAMLConfigFilename,deviceName,COMPUTING_BACKEND):
    
    class InOutputWrapper(object):
       
        def __init__(self, queue, stdout=True):
            self.queue=queue
            if stdout:
                self._stream = sys.stdout
                sys.stdout = self
            else:
                self._stream = sys.stderr
                sys.stderr = self
            self._stdout = stdout

        def write(self, text):
            self.queue.put(text)

        def __getattr__(self, name):
            return getattr(self._stream, name)

        def __del__(self):
            try:
                if self._stdout:
                    sys.stdout = self._stream
                else:
                    sys.stderr = self._stream
            except AttributeError:
                pass

    stdout = InOutputWrapper(queue,True)
  
    try:
         SSINow,SSI_BFGS=RUN_FITTING(TxConfig,
              YAMLConfigFilename,
              deviceName=deviceName,
              COMPUTING_BACKEND=COMPUTING_BACKEND)
         queue.put({'SSINow':SSINow, 'SSI_BFGS':SSI_BFGS})
    except BaseException as e:
        print('--Babel-Brain-Low-Error')
        print(traceback.format_exc())
        print(str(e))

